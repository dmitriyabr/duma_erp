"""Excel exports for parent billing account balances."""

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

MONEY_FORMAT = '#,##0.00'
HIGHLIGHT_FILL = PatternFill("solid", fgColor="DBEAFE")


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    return value


def _write_rows(ws: Worksheet, rows: list[list[Any]], start_row: int = 1) -> None:
    for row_index, row in enumerate(rows, start=start_row):
        for column_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=column_index, value=_cell_value(value))


def _style_header(ws: Worksheet, row: int, columns: int) -> None:
    fill = PatternFill("solid", fgColor="E2E8F0")
    for column in range(1, columns + 1):
        cell = ws.cell(row=row, column=column)
        cell.font = Font(bold=True)
        cell.fill = fill


def _style_label_value_row(
    ws: Worksheet,
    row: int,
    *,
    bold: bool = False,
    highlight: bool = False,
) -> None:
    for column in (1, 2):
        cell = ws.cell(row=row, column=column)
        if bold:
            cell.font = Font(bold=True, size=12 if highlight else 11)
        if highlight:
            cell.fill = HIGHLIGHT_FILL
    if isinstance(ws.cell(row=row, column=2).value, int | float):
        ws.cell(row=row, column=2).number_format = MONEY_FORMAT


def _style_money_columns(ws: Worksheet, columns: list[int], start_row: int = 1) -> None:
    for row in ws.iter_rows(min_row=start_row):
        for column in columns:
            cell = row[column - 1]
            if isinstance(cell.value, int | float):
                cell.number_format = MONEY_FORMAT


def _autosize(ws: Worksheet) -> None:
    for column_cells in ws.columns:
        width = 10
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 48))
        ws.column_dimensions[column_letter].width = width


def _money(value: Any) -> Decimal:
    if value is None:
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _invoice_term_rows(invoices: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[int | None, str], dict[str, Any]] = {}
    for invoice in invoices:
        term_id = invoice.get("term_id")
        term_name = invoice.get("term_name") or "No term / Ad-hoc"
        key = (term_id if isinstance(term_id, int) else None, term_name)
        row = grouped.setdefault(
            key,
            {
                "term_id": key[0],
                "term_name": term_name,
                "invoice_count": 0,
                "total": Decimal("0.00"),
                "paid_total": Decimal("0.00"),
                "adjustment_total": Decimal("0.00"),
                "amount_due": Decimal("0.00"),
            },
        )
        row["invoice_count"] += 1
        row["total"] += _money(invoice.get("total"))
        row["paid_total"] += _money(invoice.get("paid_total"))
        row["adjustment_total"] += _money(invoice.get("adjustment_total"))
        row["amount_due"] += _money(invoice.get("amount_due"))

    def sort_key(row: dict[str, Any]) -> tuple[int, str]:
        term_id = row.get("term_id")
        return (term_id if isinstance(term_id, int) else 999_999_999, row["term_name"])

    return sorted(grouped.values(), key=sort_key)


def build_parent_balances_xlsx(data: dict[str, Any]) -> bytes:
    """Build an all-parents balance export."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Parent Balances"

    ws.cell(1, 1, "Parent Balances")
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, "Generated on")
    ws.cell(2, 2, _cell_value(data.get("generated_on") or date.today()))
    if data.get("search"):
        ws.cell(3, 1, "Search")
        ws.cell(3, 2, data.get("search"))

    headers = [
        "Account #",
        "Parent / Family",
        "Guardian",
        "Phone",
        "Email",
        "Students",
        "Members",
        "Amount Parent Should Pay Now",
        "Overpayment / Credit",
        "Invoices Issued",
        "Paid by Parent",
        "Refunded to Parent",
        "Payment Kept by School",
        "Applied to Invoices",
        "Unused Credit",
        "Unpaid Invoice Balance",
        "Adjustments / Write-offs",
        "Last Payment",
    ]
    header_row = 5
    _write_rows(ws, [headers], header_row)
    _style_header(ws, header_row, len(headers))

    rows = data.get("rows") or []
    row_index = header_row + 1
    for row in rows:
        _write_rows(
            ws,
            [
                [
                    row.get("account_number"),
                    row.get("display_name"),
                    row.get("primary_guardian_name"),
                    row.get("primary_guardian_phone"),
                    row.get("primary_guardian_email"),
                    row.get("students"),
                    row.get("member_count"),
                    row.get("amount_to_pay_now"),
                    row.get("credit_after_debts"),
                    row.get("total_invoiced"),
                    row.get("total_payments"),
                    row.get("total_refunds"),
                    row.get("net_paid"),
                    row.get("paid_to_invoices"),
                    row.get("available_credit"),
                    row.get("outstanding_debt"),
                    row.get("invoice_adjustments"),
                    row.get("last_payment_date"),
                ]
            ],
            row_index,
        )
        row_index += 1

    summary = data.get("summary") or {}
    _write_rows(
        ws,
        [
            [
                "TOTAL",
                "",
                "",
                "",
                "",
                "",
                summary.get("account_count"),
                summary.get("amount_to_pay_now"),
                summary.get("credit_after_debts"),
                summary.get("total_invoiced"),
                summary.get("total_payments"),
                summary.get("total_refunds"),
                summary.get("net_paid"),
                summary.get("paid_to_invoices"),
                summary.get("available_credit"),
                summary.get("outstanding_debt"),
                summary.get("invoice_adjustments"),
                "",
            ]
        ],
        row_index,
    )
    for column in range(1, len(headers) + 1):
        ws.cell(row_index, column).font = Font(bold=True)

    _style_money_columns(ws, list(range(8, 18)), header_row + 1)
    _autosize(ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_parent_balance_xlsx(data: dict[str, Any]) -> bytes:
    """Build a detailed balance export for one parent billing account."""
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    account = data.get("account") or {}
    summary = data.get("summary") or {}
    amount_to_pay = summary.get("amount_to_pay_now")
    credit_after_debts = summary.get("credit_after_debts")
    summary_rows = [
        ["Parent Balance"],
        ["Generated on", data.get("generated_on") or date.today()],
        ["Account #", account.get("account_number")],
        ["Parent / Family", account.get("display_name")],
        ["Guardian", account.get("primary_guardian_name")],
        ["Phone", account.get("primary_guardian_phone")],
        ["Email", account.get("primary_guardian_email")],
        ["Students", account.get("students")],
        [],
        ["Simple Summary", ""],
        ["Amount parent should pay now", amount_to_pay],
        ["Paid by parent", summary.get("total_payments")],
        ["Refunded to parent", summary.get("total_refunds")],
        ["Payment kept by school", summary.get("net_paid")],
        ["Overpayment / credit after unpaid invoices", credit_after_debts],
        [],
        ["How this was calculated", ""],
        ["Invoices issued to this family", summary.get("total_invoiced")],
        ["Already applied to invoices", summary.get("paid_to_invoices")],
        ["Unused credit before unpaid invoices", summary.get("available_credit")],
        ["Unpaid invoice balance", summary.get("outstanding_debt")],
        ["Adjustments / write-offs", summary.get("invoice_adjustments")],
        [],
        [
            "Formula",
            "Amount to pay now = unpaid invoice balance - unused credit, never below zero.",
        ],
    ]
    _write_rows(summary_ws, summary_rows)
    summary_ws.cell(1, 1).font = Font(bold=True, size=14)
    _style_label_value_row(summary_ws, 10, bold=True)
    _style_label_value_row(summary_ws, 11, bold=True, highlight=True)
    _style_label_value_row(summary_ws, 12, bold=True)
    _style_label_value_row(summary_ws, 13, bold=True)
    _style_label_value_row(summary_ws, 14, bold=True)
    _style_label_value_row(summary_ws, 15, bold=True)
    _style_label_value_row(summary_ws, 17, bold=True)
    _style_money_columns(summary_ws, [2], 11)
    _autosize(summary_ws)

    invoices = data.get("invoices") or []
    terms_ws = wb.create_sheet("Invoices by Term")
    term_headers = [
        "Term",
        "Invoices",
        "Total Invoiced",
        "Paid",
        "Adjustments",
        "Amount Due",
    ]
    _write_rows(terms_ws, [term_headers], 1)
    _style_header(terms_ws, 1, len(term_headers))
    for index, term_row in enumerate(_invoice_term_rows(invoices), start=2):
        _write_rows(
            terms_ws,
            [[
                term_row.get("term_name"),
                term_row.get("invoice_count"),
                term_row.get("total"),
                term_row.get("paid_total"),
                term_row.get("adjustment_total"),
                term_row.get("amount_due"),
            ]],
            index,
        )
    _style_money_columns(terms_ws, [3, 4, 5, 6], 2)
    _autosize(terms_ws)

    students_ws = wb.create_sheet("Students")
    student_headers = ["Student #", "Student", "Grade", "Guardian", "Phone", "Status"]
    _write_rows(students_ws, [student_headers], 1)
    _style_header(students_ws, 1, len(student_headers))
    for index, student in enumerate(data.get("students") or [], start=2):
        _write_rows(
            students_ws,
            [[
                student.get("student_number"),
                student.get("student_name"),
                student.get("grade_name"),
                student.get("guardian_name"),
                student.get("guardian_phone"),
                student.get("status"),
            ]],
            index,
        )
    _autosize(students_ws)

    invoices_ws = wb.create_sheet("Invoices")
    invoice_headers = [
        "Term",
        "Invoice #",
        "Student",
        "Type",
        "Status",
        "Issue Date",
        "Due Date",
        "Total",
        "Paid",
        "Adjustments",
        "Amount Due",
    ]
    _write_rows(invoices_ws, [invoice_headers], 1)
    _style_header(invoices_ws, 1, len(invoice_headers))
    for index, invoice in enumerate(invoices, start=2):
        _write_rows(
            invoices_ws,
            [[
                invoice.get("term_name") or "No term / Ad-hoc",
                invoice.get("invoice_number"),
                invoice.get("student_name"),
                invoice.get("invoice_type"),
                invoice.get("status"),
                invoice.get("issue_date"),
                invoice.get("due_date"),
                invoice.get("total"),
                invoice.get("paid_total"),
                invoice.get("adjustment_total"),
                invoice.get("amount_due"),
            ]],
            index,
        )
    _style_money_columns(invoices_ws, [8, 9, 10, 11], 2)
    _autosize(invoices_ws)

    payments_ws = wb.create_sheet("Payments")
    payment_headers = [
        "Date",
        "Payment #",
        "Receipt #",
        "Student",
        "Method",
        "Reference",
        "Status",
        "Amount",
        "Refunded",
        "Net Amount",
    ]
    _write_rows(payments_ws, [payment_headers], 1)
    _style_header(payments_ws, 1, len(payment_headers))
    for index, payment in enumerate(data.get("payments") or [], start=2):
        _write_rows(
            payments_ws,
            [[
                payment.get("payment_date"),
                payment.get("payment_number"),
                payment.get("receipt_number"),
                payment.get("student_name"),
                payment.get("payment_method"),
                payment.get("reference"),
                payment.get("status"),
                payment.get("amount"),
                payment.get("refunded_amount"),
                payment.get("net_amount"),
            ]],
            index,
        )
    _style_money_columns(payments_ws, [8, 9, 10], 2)
    _autosize(payments_ws)

    refunds_ws = wb.create_sheet("Refunds")
    refund_headers = [
        "Date",
        "Refund #",
        "Method",
        "Reference",
        "Reason",
        "Amount",
    ]
    _write_rows(refunds_ws, [refund_headers], 1)
    _style_header(refunds_ws, 1, len(refund_headers))
    for index, refund in enumerate(data.get("refunds") or [], start=2):
        _write_rows(
            refunds_ws,
            [[
                refund.get("refund_date"),
                refund.get("refund_number"),
                refund.get("refund_method"),
                refund.get("reference_number"),
                refund.get("reason"),
                refund.get("amount"),
            ]],
            index,
        )
    _style_money_columns(refunds_ws, [6], 2)
    _autosize(refunds_ws)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
