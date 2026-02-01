"""Export report data to Excel (XLSX)."""

from decimal import Decimal
from io import BytesIO
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Font


def _cell_value(v: Any) -> Any:
    """Convert value for Excel (Decimal -> float, date stays)."""
    if v is None:
        return None
    if isinstance(v, Decimal):
        return float(v)
    return v


def _write_table(ws: Any, rows: list[list[Any]], start_row: int = 1) -> None:
    """Write list of rows to sheet starting at start_row."""
    for i, row in enumerate(rows, start=start_row):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=_cell_value(val))


def export_aged_receivables(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Aged Receivables"
    as_at = data.get("as_at_date")
    ws.cell(1, 1, f"Aged Receivables as at {as_at}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Student ID", "Student Name", "Total", "Current (0-30)", "31-60", "61-90", "90+", "Last Payment"]
    _write_table(ws, [headers], 3)
    for c in range(1, len(headers) + 1):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[
            r.get("student_id"), r.get("student_name"), r.get("total"), r.get("current"),
            r.get("bucket_31_60"), r.get("bucket_61_90"), r.get("bucket_90_plus"),
            r.get("last_payment_date"),
        ]], row)
        row += 1
    s = data.get("summary", {})
    _write_table(ws, [["TOTAL", "", s.get("total"), s.get("current"), s.get("bucket_31_60"),
                      s.get("bucket_61_90"), s.get("bucket_90_plus"), ""]], row)
    for c in range(1, 4):
        ws.cell(row, c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_student_fees(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Fees"
    term = data.get("term_display_name", "")
    ws.cell(1, 1, f"Student Fees by Term: {term}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Grade", "Students", "Total Invoiced", "Total Paid", "Balance", "Rate %"]
    _write_table(ws, [headers], 3)
    for c in range(1, len(headers) + 1):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[
            r.get("grade_name"), r.get("students_count"), r.get("total_invoiced"),
            r.get("total_paid"), r.get("balance"), r.get("rate_percent"),
        ]], row)
        row += 1
    s = data.get("summary", {})
    _write_table(ws, [["TOTAL", s.get("students_count"), s.get("total_invoiced"),
                      s.get("total_paid"), s.get("balance"), s.get("rate_percent")]], row)
    for c in range(1, 7):
        ws.cell(row, c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_profit_loss_fixed(data: dict) -> bytes:
    """Fixed version: use dict.get for monthly values."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    ws.cell(1, 1, f"Profit & Loss: {date_from} to {date_to}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    months = data.get("months") or []
    if months:
        header = ["Line"] + [m.replace("-", " ") for m in months] + ["Total (KES)"]
    else:
        header = ["Line", "Amount (KES)"]
    _write_table(ws, [header], 3)
    for c in range(1, len(header) + 1):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    ws.cell(row, 1, "REVENUE").font = Font(bold=True)
    row += 1
    for r in data.get("revenue_lines", []):
        monthly = r.get("monthly") or {}
        line = [r.get("label")] + [_cell_value(monthly.get(m)) for m in months] + [r.get("amount")]
        if not months:
            line = [r.get("label"), r.get("amount")]
        _write_table(ws, [line], row)
        row += 1
    gr = data.get("gross_revenue")
    gr_m = data.get("gross_revenue_monthly") or {}
    line = ["Gross Revenue"] + [_cell_value(gr_m.get(m)) for m in months] + [gr]
    if not months:
        line = ["Gross Revenue", gr]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    td = data.get("total_discounts")
    td_m = data.get("total_discounts_monthly") or {}
    line = ["Less: Discounts"] + [_cell_value(-float(td_m.get(m) or 0)) for m in months] + [-float(td or 0)]
    if not months:
        line = ["Less: Discounts", -float(td or 0)]
    _write_table(ws, [line], row)
    row += 1
    nr_m = data.get("net_revenue_monthly") or {}
    line = ["Net Revenue"] + [_cell_value(nr_m.get(m)) for m in months] + [data.get("net_revenue")]
    if not months:
        line = ["Net Revenue", data.get("net_revenue")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    row += 2
    ws.cell(row, 1, "EXPENSES").font = Font(bold=True)
    row += 1
    for r in data.get("expense_lines", []):
        monthly = r.get("monthly") or {}
        line = [r.get("label")] + [_cell_value(monthly.get(m)) for m in months] + [r.get("amount")]
        if not months:
            line = [r.get("label"), r.get("amount")]
        _write_table(ws, [line], row)
        row += 1
    te_m = data.get("total_expenses_monthly") or {}
    line = ["Total Expenses"] + [_cell_value(te_m.get(m)) for m in months] + [data.get("total_expenses")]
    if not months:
        line = ["Total Expenses", data.get("total_expenses")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    row += 2
    np_m = data.get("net_profit_monthly") or {}
    line = ["NET PROFIT"] + [_cell_value(np_m.get(m)) for m in months] + [data.get("net_profit")]
    if not months:
        line = ["NET PROFIT", data.get("net_profit")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_cash_flow(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"
    date_from = data.get("date_from")
    date_to = data.get("date_to")
    ws.cell(1, 1, f"Cash Flow: {date_from} to {date_to}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    months = data.get("months") or []
    if months:
        header = ["Line"] + [m.replace("-", " ") for m in months] + ["Total (KES)"]
    else:
        header = ["Line", "Amount (KES)"]
    _write_table(ws, [header], 3)
    for c in range(1, len(header) + 1):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    _write_table(ws, [["Opening Balance"] + ([""] * len(months) if months else []) + [data.get("opening_balance")]], row)
    row += 1
    ws.cell(row, 1, "INFLOWS").font = Font(bold=True)
    row += 1
    for r in data.get("inflow_lines", []):
        monthly = r.get("monthly") or {}
        line = [r.get("label")] + [_cell_value(monthly.get(m)) for m in months] + [r.get("amount")]
        if not months:
            line = [r.get("label"), r.get("amount")]
        _write_table(ws, [line], row)
        row += 1
    ti_m = data.get("total_inflows_monthly") or {}
    line = ["Total Inflows"] + [_cell_value(ti_m.get(m)) for m in months] + [data.get("total_inflows")]
    if not months:
        line = ["Total Inflows", data.get("total_inflows")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    row += 2
    ws.cell(row, 1, "OUTFLOWS").font = Font(bold=True)
    row += 1
    for r in data.get("outflow_lines", []):
        monthly = r.get("monthly") or {}
        line = [r.get("label")] + [_cell_value(monthly.get(m)) for m in months] + [r.get("amount")]
        if not months:
            line = [r.get("label"), r.get("amount")]
        _write_table(ws, [line], row)
        row += 1
    to_m = data.get("total_outflows_monthly") or {}
    line = ["Total Outflows"] + [_cell_value(to_m.get(m)) for m in months] + [data.get("total_outflows")]
    if not months:
        line = ["Total Outflows", data.get("total_outflows")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    row += 2
    _write_table(ws, [["Net Cash Flow"] + ([""] * len(months) if months else []) + [data.get("net_cash_flow")]], row)
    ws.cell(row, 1).font = Font(bold=True)
    row += 1
    cb_m = data.get("closing_balance_monthly") or {}
    line = ["Closing Balance"] + [_cell_value(cb_m.get(m)) for m in months] + [data.get("closing_balance")]
    if not months:
        line = ["Closing Balance", data.get("closing_balance")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_balance_sheet(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    as_at = data.get("as_at_date")
    ws.cell(1, 1, f"Balance Sheet as at {as_at}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    months = data.get("months") or []
    if months:
        header = ["Line"] + [m.replace("-", " ") for m in months] + ["Total (KES)"]
    else:
        header = ["Line", "Amount (KES)"]
    _write_table(ws, [header], 3)
    for c in range(1, len(header) + 1):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    ws.cell(row, 1, "ASSETS").font = Font(bold=True)
    row += 1
    for r in data.get("asset_lines", []):
        monthly = r.get("monthly") or {}
        line = [r.get("label")] + [_cell_value(monthly.get(m)) for m in months] + [r.get("amount")]
        if not months:
            line = [r.get("label"), r.get("amount")]
        _write_table(ws, [line], row)
        row += 1
    ta_m = data.get("total_assets_monthly") or {}
    line = ["Total Assets"] + [_cell_value(ta_m.get(m)) for m in months] + [data.get("total_assets")]
    if not months:
        line = ["Total Assets", data.get("total_assets")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    row += 2
    ws.cell(row, 1, "LIABILITIES").font = Font(bold=True)
    row += 1
    for r in data.get("liability_lines", []):
        monthly = r.get("monthly") or {}
        line = [r.get("label")] + [_cell_value(monthly.get(m)) for m in months] + [r.get("amount")]
        if not months:
            line = [r.get("label"), r.get("amount")]
        _write_table(ws, [line], row)
        row += 1
    tl_m = data.get("total_liabilities_monthly") or {}
    line = ["Total Liabilities"] + [_cell_value(tl_m.get(m)) for m in months] + [data.get("total_liabilities")]
    if not months:
        line = ["Total Liabilities", data.get("total_liabilities")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    row += 2
    ne_m = data.get("net_equity_monthly") or {}
    line = ["Net Equity"] + [_cell_value(ne_m.get(m)) for m in months] + [data.get("net_equity")]
    if not months:
        line = ["Net Equity", data.get("net_equity")]
    _write_table(ws, [line], row)
    ws.cell(row, 1).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_collection_rate(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Collection Rate"
    ws.cell(1, 1, "Collection Rate Trend")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Month", "Total Invoiced", "Total Paid", "Rate %"]
    _write_table(ws, [headers], 3)
    for c in range(1, 5):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("label"), r.get("total_invoiced"), r.get("total_paid"), r.get("rate_percent")]], row)
        row += 1
    ws.cell(row, 1, "Average").font = Font(bold=True)
    ws.cell(row, 4, data.get("average_rate_percent")).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_discount_analysis(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Discount Analysis"
    ws.cell(1, 1, f"Discount Analysis: {data.get('date_from')} to {data.get('date_to')}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Reason", "Students", "Total Amount", "Avg per Student", "% of Revenue"]
    _write_table(ws, [headers], 3)
    for c in range(1, 6):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("reason_name"), r.get("students_count"), r.get("total_amount"),
                          r.get("avg_per_student"), r.get("percent_of_revenue")]], row)
        row += 1
    s = data.get("summary", {})
    _write_table(ws, [["TOTAL", s.get("students_count"), s.get("total_discount_amount"), "", s.get("percent_of_revenue")]], row)
    for c in range(1, 6):
        ws.cell(row, c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_top_debtors(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top Debtors"
    ws.cell(1, 1, f"Top Debtors as at {data.get('as_at_date')}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Student", "Grade", "Total Debt", "Invoice Count", "Oldest Due"]
    _write_table(ws, [headers], 3)
    for c in range(1, 6):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("student_name"), r.get("grade_name"), r.get("total_debt"),
                          r.get("invoice_count"), r.get("oldest_due_date")]], row)
        row += 1
    _write_table(ws, [["TOTAL DEBT", "", data.get("total_debt"), "", ""]], row)
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 3).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_procurement_summary(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Procurement Summary"
    ws.cell(1, 1, f"Procurement Summary: {data.get('date_from')} to {data.get('date_to')}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Supplier", "PO Count", "Total", "Paid", "Outstanding", "Status"]
    _write_table(ws, [headers], 3)
    for c in range(1, 7):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("supplier_name"), r.get("po_count"), r.get("total_amount"),
                          r.get("paid"), r.get("outstanding"), r.get("status")]], row)
        row += 1
    _write_table(ws, [["TOTAL", data.get("total_po_count"), data.get("total_amount"),
                      data.get("total_paid"), data.get("total_outstanding"), ""]], row)
    for c in range(1, 7):
        ws.cell(row, c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_inventory_valuation(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Valuation"
    ws.cell(1, 1, f"Inventory Valuation as at {data.get('as_at_date')}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Category", "Items", "Quantity", "Unit Cost Avg", "Total Value", "Turnover"]
    _write_table(ws, [headers], 3)
    for c in range(1, 7):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("category_name"), r.get("items_count"), r.get("quantity"),
                          r.get("unit_cost_avg"), r.get("total_value"), r.get("turnover")]], row)
        row += 1
    _write_table(ws, [["TOTAL", data.get("total_items"), data.get("total_quantity"), "",
                      data.get("total_value"), ""]], row)
    for c in range(1, 7):
        ws.cell(row, c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_low_stock_alert(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Low Stock Alert"
    ws.cell(1, 1, "Low Stock Alert")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Item", "SKU", "Current", "Min Level", "Status", "Suggested Order"]
    _write_table(ws, [headers], 3)
    for c in range(1, 7):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("item_name"), r.get("sku_code"), r.get("current"), r.get("min_level"),
                          r.get("status"), r.get("suggested_order")]], row)
        row += 1
    ws.cell(row, 1, f"Total low/out: {data.get('total_low_count', 0)}").font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_stock_movement(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Movement"
    ws.cell(1, 1, f"Stock Movement: {data.get('date_from')} to {data.get('date_to')}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Date", "Type", "Item", "Quantity", "Ref", "By", "Balance After"]
    _write_table(ws, [headers], 3)
    for c in range(1, 8):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("movement_date"), r.get("movement_type"), r.get("item_name"),
                          r.get("quantity"), r.get("ref_display"), r.get("created_by_name"),
                          r.get("balance_after")]], row)
        row += 1
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_compensation_summary(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compensation Summary"
    ws.cell(1, 1, f"Compensation Summary: {data.get('date_from')} to {data.get('date_to')}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Employee", "Claims", "Total", "Approved", "Paid", "Pending"]
    _write_table(ws, [headers], 3)
    for c in range(1, 7):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("employee_name"), r.get("claims_count"), r.get("total_amount"),
                          r.get("approved_amount"), r.get("paid_amount"), r.get("pending_amount")]], row)
        row += 1
    s = data.get("summary", {})
    _write_table(ws, [["TOTAL", s.get("total_claims"), s.get("total_amount"), s.get("total_approved"),
                      s.get("total_paid"), s.get("total_pending")]], row)
    for c in range(1, 7):
        ws.cell(row, c).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_expense_claims_by_category(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Claims by Category"
    ws.cell(1, 1, f"Expense Claims by Category: {data.get('date_from')} to {data.get('date_to')}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Category", "Amount", "Claims Count", "% of Total"]
    _write_table(ws, [headers], 3)
    for c in range(1, 5):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("purpose_name"), r.get("amount"), r.get("claims_count"), r.get("percent_of_total")]], row)
        row += 1
    _write_table(ws, [["TOTAL", data.get("total_amount"), "", ""]], row)
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_revenue_trend(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue Trend"
    ws.cell(1, 1, "Revenue per Student Trend")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Year", "Total Revenue", "Students", "Avg per Student"]
    _write_table(ws, [headers], 3)
    for c in range(1, 5):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("label"), r.get("total_revenue"), r.get("students_count"), r.get("avg_revenue_per_student")]], row)
        row += 1
    ws.cell(row, 1, "Growth %").font = Font(bold=True)
    ws.cell(row, 2, data.get("growth_percent")).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_payment_method_distribution(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Method Distribution"
    ws.cell(1, 1, f"Payment Method Distribution: {data.get('date_from')} to {data.get('date_to')}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Method", "Amount", "% of Total"]
    _write_table(ws, [headers], 3)
    for c in range(1, 4):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for r in data.get("rows", []):
        _write_table(ws, [[r.get("label"), r.get("amount"), r.get("percent_of_total")]], row)
        row += 1
    _write_table(ws, [["TOTAL", data.get("total_amount"), ""]], row)
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2).font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_term_comparison(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Term Comparison"
    t1 = data.get("term1_display_name", "")
    t2 = data.get("term2_display_name", "")
    ws.cell(1, 1, f"Term Comparison: {t1} vs {t2}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    headers = ["Metric", "Term 1", "Term 2", "Change", "Change %"]
    _write_table(ws, [headers], 3)
    for c in range(1, 6):
        ws.cell(3, c).font = Font(bold=True)
    row = 4
    for m in data.get("metrics", []):
        _write_table(ws, [[m.get("name"), m.get("term1_value"), m.get("term2_value"), m.get("change_abs"), m.get("change_percent")]], row)
        row += 1
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_kpis(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "KPIs"
    period = data.get("term_display_name") or str(data.get("year", ""))
    ws.cell(1, 1, f"KPIs & Metrics: {period}")
    ws.cell(1, 1).font = Font(bold=True, size=12)
    row = 3
    metrics = [
        ("Active Students", data.get("active_students_count")),
        ("Total Revenue", data.get("total_revenue")),
        ("Total Invoiced", data.get("total_invoiced")),
        ("Collection Rate %", data.get("collection_rate_percent")),
        ("Total Expenses", data.get("total_expenses")),
        ("Student Debt", data.get("student_debt")),
        ("Supplier Debt", data.get("supplier_debt")),
        ("Pending Claims", data.get("pending_claims_amount")),
    ]
    for name, val in metrics:
        ws.cell(row, 1, name)
        ws.cell(row, 2, _cell_value(val))
        row += 1
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


_EXPORTERS: dict[str, Callable[[dict], bytes]] = {
    "aged_receivables": export_aged_receivables,
    "student_fees": export_student_fees,
    "profit_loss": export_profit_loss_fixed,
    "cash_flow": export_cash_flow,
    "balance_sheet": export_balance_sheet,
    "collection_rate": export_collection_rate,
    "discount_analysis": export_discount_analysis,
    "top_debtors": export_top_debtors,
    "procurement_summary": export_procurement_summary,
    "inventory_valuation": export_inventory_valuation,
    "low_stock_alert": export_low_stock_alert,
    "stock_movement": export_stock_movement,
    "compensation_summary": export_compensation_summary,
    "expense_claims_by_category": export_expense_claims_by_category,
    "revenue_trend": export_revenue_trend,
    "payment_method_distribution": export_payment_method_distribution,
    "term_comparison": export_term_comparison,
    "kpis": export_kpis,
}


def build_report_xlsx(report_key: str, data: dict) -> bytes:
    """Build XLSX bytes for a report. report_key e.g. 'aged_receivables', 'profit_loss'."""
    fn = _EXPORTERS.get(report_key)
    if not fn:
        raise ValueError(f"Unknown report for Excel: {report_key}")
    return fn(data)
